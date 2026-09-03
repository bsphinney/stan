#!/usr/bin/env python3
"""export_rerun_package.py — the three deliverables for a submission's reruns.

Reads the QC scan produced by `qc_scan.py` (frame counts + TIC read straight
from each .d/analysis.tdf) and writes:

  1. <PROT_xxxx>_reruns.tsv        tab-delimited, one row per rerun candidate
  2. <PROT_xxxx>_rerun_queue.xlsx  a HyStar SampleTable matching the format of
                                   the original exported queue
  3. <PROT_xxxx>_platemap.pdf      one page per plate, wells shaded by TIC,
                                   rerun candidates ringed

**The STAN dashboard is the authority** on which acquired wells are suspect.
Pass `--dashboard` and its `needs_rerun` decides: it scores six metrics per
plate against a robust z, where this script only ever measured TIC, so a second
local opinion would just be a second list that disagrees.

  flagged     STAN flagged it. The reason it gave is carried through verbatim.
  no_data     the .d holds no analysis.tdf — the acquisition produced nothing.

The no-data class is detected HERE and unioned in, always. A run that produced
no file never got a sample_health row, so it cannot appear in `needs_rerun` at
all — the dashboard is structurally blind to exactly the wells that most
certainly need re-running. A sample that makes no data file is an automatic
rerun, so it is never gated on STAN's verdict.

Blanks are excluded from every class, whatever the dashboard says: a wash is
SUPPOSED to be empty, and re-injecting one buys nothing.

Without `--dashboard` the script falls back to a local rule (TIC under
`--blank-frac` of the plate median, default 25%). That is the offline path; it
will not agree with the dashboard and is not meant to.

    python3 export_rerun_package.py --scan qc_scan.json --submission 793 \
        --dashboard https://ucd.stan-proteomics.org --dash-token <token> \
        --run-date 20260902 --out-dir ./exports

    python3 export_rerun_package.py --scan qc_scan.json --submission 793   # offline
"""

from __future__ import annotations

import argparse
import json
import re
import statistics as st
import urllib.parse
import urllib.request
from pathlib import Path

# ── the exported-queue format, taken from the original PROTIFI export ─────────
QUEUE_SHEET = "SampleTable"
QUEUE_COLUMNS = [
    "CheckToRun", "Vial", "Sample ID", "Separation Method", "MS Method",
    "Status", "Volume [µl]", "Data Path", "Result Path", "Sample Comment",
    "Start Date", "End Date",
]
SEP_METHOD = r"D:\Methods\EvoSepLCmeth\100spd.m?HyStar_LC"
MS_METHOD = (r"D:\Methods\MSmeth\ela\wBPS_11Ian24"
             r"\DIA_11x3-k07t13Ra85.m?OtofImpacTEMControl")
DATA_PATH = r"D:\Data\Aug26"
QUEUE_FONT = ("Tahoma", 10)
QUEUE_WIDTHS = {"C": 32.86, "D": 33.57, "E": 67.71, "F": 23.71}

# ── palette (dataviz: sequential = one hue light→dark; status is reserved) ────
SEQ_HUE = "#2a78d6"          # blue, the default sequential hue
CRITICAL = "#d03b3b"         # status: rerun / no data
WARNING = "#fab219"          # status: at-blank
SURFACE = "#fcfcfb"
INK = "#0b0b0b"
INK_MUTED = "#52514e"
NEUTRAL = "#f0efec"

ROWS = "ABCDEFGH"
COLS = list(range(1, 13))


def classify(name: str) -> str:
    if "Hel" in name or "HeL" in name:
        return "HeLa"
    if "Blank" in name.lower() or "blank" in name.lower():
        return "Blank"
    return "sample"


# Anchored on the well token, which every acquisition carries, rather than on
# the submission number, which some trays were acquired without. Parsing from
# the submission end silently fell through on those and emitted whole filenames
# where a sample name belongs — wrong in the queue, where it becomes the name
# the instrument writes.
_PARTS = re.compile(
    r"^(?P<date>\d{8})_"
    r"(?:(?P<sub>\d{2,4})_)?"            # submission number: optional
    r"(?P<method>[^_]+)_"
    r"(?P<samp>.+?)_"
    r"S\d+-[A-H]\d{1,2}_\d+_\d+\.d$"
)


def parse_run(run: str) -> dict | None:
    m = _PARTS.match(run)
    return m.groupdict() if m else None


def sample_label(run: str) -> str:
    """`20260828_793_100spd_COH-35_S5-C5_1_24155.d` -> `COH-35`."""
    p = parse_run(run)
    return p["samp"] if p else run.removesuffix(".d")


def queue_stem(run: str, run_date: str, submission: str) -> str:
    """The Sample ID the instrument is given: date_sub_method_sample.

    The acquisition software appends `_S<tray>-<well>_1_<counter>` itself, so
    the queue must NOT carry them — that is why the original export's Sample ID
    stops at the sample name.
    """
    p = parse_run(run)
    if not p:
        return run.removesuffix(".d")
    sub = p["sub"] or str(int(submission))
    return f"{run_date}_{sub}_{p['method']}_{p['samp']}"


def fetch_dashboard_reruns(base: str, q: str, token: str | None,
                           instrument: str) -> dict[str, str]:
    """STAN's own `needs_rerun`, keyed by injection number.

    The dashboard is the authority on WHICH wells are suspect: it scores six
    metrics per plate against a robust z, where this script only ever looked at
    TIC. Keeping a second opinion here just means two lists that disagree.

    Keyed by injection rather than by name because a run renamed after it was
    first processed appears twice — once under each name — and both rows point
    at one physical acquisition.
    """
    url = (f"{base.rstrip('/')}/api/ht/submission?q={urllib.parse.quote(q)}"
           f"&instrument={urllib.parse.quote(instrument)}&metric=ms1_total_tic")
    if token:
        url += f"&token={urllib.parse.quote(token)}"
    with urllib.request.urlopen(url, timeout=60) as resp:
        payload = json.loads(resp.read())
    out: dict[str, str] = {}
    for r in payload.get("needs_rerun", []):
        name = r.get("run_name") or ""
        m = re.search(r"_(\d+)\.d$", name)
        if not m:
            continue
        # Never carry a control into the list, whatever the dashboard says —
        # an older deployment scores blanks as samples.
        if classify(name) == "Blank":
            continue
        why = r.get("outlier_reasons") or r.get("reasons") or []
        if isinstance(why, list) and why:
            why = "; ".join(str(x.get("label", x) if isinstance(x, dict) else x)
                            for x in why[:3])
        out[m.group(1)] = str(why) if why else "flagged by STAN"
    return out


def read_queue(paths: list[str]) -> dict[str, str]:
    """well -> intended sample name, from the submitted HyStar queue(s).

    The queue is the only record that a well was supposed to run at all. A well
    that produced no directory is invisible to a scan of directories and has no
    sample_health row, so neither this script nor STAN can see it — it simply
    is not there. Reconciling against the queue is what turns "nothing found"
    into "nothing ran", which is a rerun.
    """
    import openpyxl
    out: dict[str, str] = {}
    for p in paths:
        ws = openpyxl.load_workbook(p)[QUEUE_SHEET]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vial, sid = row[1], row[2]
            if not vial or "-" not in str(vial):
                continue
            well = str(vial).split("-", 1)[1]
            out[well] = str(sid or "").split("_")[-1]
    return out


def find_missing_wells(rows: list[dict], queue: dict[str, str]) -> list[dict]:
    """Queued wells with no directory on disk at all."""
    seen = {r["well"] for r in rows}
    out = []
    for well, samp in sorted(queue.items()):
        if well in seen or classify(samp) == "Blank":
            continue
        plate = rows[0]["plate"] if rows else "?"
        out.append({"run": f"{samp}.d", "plate": plate, "well": well,
                    "inj": None, "reason": "never_ran",
                    "detail": "queued but no directory was ever written"})
    return out


def find_reruns(rows: list[dict], frac: float,
                dashboard: dict[str, str] | None = None) -> list[dict]:
    """Rerun candidates: sample wells only.

    With `dashboard`, STAN's verdict decides which acquired wells are suspect
    and the local TIC rule is not applied. The no-data class is still detected
    here and unioned in — a `.d` holding no analysis.tdf never produced a
    sample_health row, so it cannot appear in `needs_rerun` at all. Those are
    the most certain reruns of the lot, and taking the dashboard as the whole
    answer would silently drop them.

    Blanks are excluded from every class. A blank is a wash — it is SUPPOSED
    to be empty, so judging it against a sample median flags it for being
    exactly what it should be, and re-injecting one buys nothing. (A blank
    carrying real signal is a carryover problem, which is a different report
    and is not fixed by running it again.)
    """
    out = []
    for r in rows:
        if "error" in r and classify(r["run"]) != "Blank":
            out.append({**r, "reason": "no_data",
                        "detail": "no analysis.tdf — acquisition produced nothing"})
    ok = [r for r in rows if "error" not in r]
    for r in ok:
        r["cls"] = classify(r["run"])
    plates = {p: [r for r in ok if r["plate"] == p and r["cls"] == "sample"]
              for p in {r["plate"] for r in ok}}
    medians = {p: st.median([r["tic"] for r in g]) for p, g in plates.items() if g}

    if dashboard is not None:
        for plate, grp in plates.items():
            for r in grp:
                why = dashboard.get(str(r.get("inj")))
                if why:
                    pct = r["tic"] / medians[plate] * 100
                    out.append({**r, "reason": "flagged",
                                "detail": f"{why} — TIC {pct:.1f}% of {plate} median"})
    else:
        for plate, grp in plates.items():
            med = medians[plate]
            for r in grp:
                if r["tic"] < frac * med:
                    out.append({**r, "reason": "at_blank",
                                "detail": f"TIC {r['tic']/med*100:.1f}% of {plate} median"})
    out.sort(key=lambda r: (r["plate"], r["well"][0], int(r["well"][1:])))
    return out


def write_tsv(reruns: list[dict], path: Path) -> None:
    cols = ["plate", "well", "injection", "sample", "reason", "detail",
            "ms2_frames", "tic", "run"]
    with path.open("w") as fh:
        fh.write("\t".join(cols) + "\n")
        for r in reruns:
            fh.write("\t".join(str(x) for x in [
                r.get("plate", ""), r.get("well", ""), r.get("inj", ""),
                sample_label(r["run"]), r["reason"], r["detail"],
                r.get("ms2", ""), r.get("tic", ""), r["run"],
            ]) + "\n")


def write_queue(reruns: list[dict], path: Path, run_date: str,
                submission: str) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = QUEUE_SHEET
    ws.append(QUEUE_COLUMNS)

    # Fresh plate, filled column-major (A1,B1..H1,A2..) exactly as the original.
    slots = [f"S1-{r}{c}" for c in COLS for r in ROWS]
    for i, r in enumerate(reruns):
        if i >= len(slots):
            break
        ws.append([
            "True", slots[i], queue_stem(r["run"], run_date, submission),
            SEP_METHOD, MS_METHOD, None, 0, DATA_PATH, None,
            f"rerun of {r.get('plate','')}-{r.get('well','')} ({r['reason']})",
            None, None,
        ])

    f = Font(name=QUEUE_FONT[0], size=QUEUE_FONT[1])
    for row in ws.iter_rows():
        for c in row:
            c.font = f
    for col, w in QUEUE_WIDTHS.items():
        ws.column_dimensions[col].width = w
    wb.save(path)


def write_platemap(rows: list[dict], reruns: list[dict], path: Path,
                   submission: str) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.colors import LinearSegmentedColormap, Normalize
    from matplotlib.ticker import FuncFormatter
    from matplotlib.patches import Rectangle

    # Sequential: one hue, light -> dark. Never a rainbow.
    cmap = LinearSegmentedColormap.from_list("seq", ["#eef4fc", SEQ_HUE, "#12325c"])
    flagged = {(r.get("plate"), r.get("well")): r["reason"] for r in reruns}
    by_plate: dict[str, dict] = {}
    for r in rows:
        by_plate.setdefault(r["plate"], {})[r["well"]] = r

    with PdfPages(path) as pdf:
        for plate in sorted(by_plate, reverse=True):
            wells = by_plate[plate]
            tics = [w["tic"] for w in wells.values()
                    if "error" not in w and classify(w["run"]) == "sample"]
            norm = Normalize(vmin=min(tics), vmax=max(tics)) if tics else Normalize(0, 1)

            fig, ax = plt.subplots(figsize=(13.5, 8.0))
            fig.patch.set_facecolor(SURFACE)
            ax.set_facecolor(SURFACE)

            for ri, rl in enumerate(ROWS):
                for c in COLS:
                    well = f"{rl}{c}"
                    w = wells.get(well)
                    x, y = c - 0.5, len(ROWS) - ri - 0.5
                    if w is None:
                        ax.add_patch(Rectangle((x, y), .92, .92, facecolor=NEUTRAL,
                                               edgecolor="#d8d6d1", lw=.8))
                        continue
                    missing = "error" in w
                    kind = classify(w["run"])
                    if missing:
                        fc, hatch = NEUTRAL, "///"
                    else:
                        fc, hatch = cmap(norm(w["tic"])), None
                    ax.add_patch(Rectangle((x, y), .92, .92, facecolor=fc,
                                           edgecolor="#c9c7c2", lw=.8, hatch=hatch))
                    reason = flagged.get((plate, well))
                    if reason:
                        ax.add_patch(Rectangle(
                            (x - .03, y - .03), .98, .98, fill=False,
                            edgecolor=CRITICAL if reason == "no_data" else WARNING,
                            lw=2.6, zorder=5))
                    # Text wears ink tokens, never the mark's own colour.
                    lum = 0 if missing else norm(w["tic"])
                    txt = "#ffffff" if (not missing and lum > .55) else INK
                    ax.text(x + .46, y + .60, well, ha="center", va="center",
                            fontsize=7.5, color=txt, weight="bold")
                    lbl = sample_label(w["run"])
                    ax.text(x + .46, y + .38, lbl[:11], ha="center", va="center",
                            fontsize=5.8, color=txt)
                    if not missing:
                        ax.text(x + .46, y + .17, f"{w['tic']/1e9:.1f}",
                                ha="center", va="center", fontsize=5.4, color=txt)
                    else:
                        ax.text(x + .46, y + .17, "NO DATA", ha="center",
                                va="center", fontsize=5.2, color=CRITICAL,
                                weight="bold")

            ax.set_xlim(0.3, 12.65)
            ax.set_ylim(-0.35, 8.3)
            ax.set_xticks([c - 0.04 for c in COLS])
            ax.set_xticklabels(COLS, fontsize=9, color=INK_MUTED)
            ax.set_yticks([len(ROWS) - i - 0.04 for i in range(len(ROWS))])
            ax.set_yticklabels(list(ROWS), fontsize=9, color=INK_MUTED)
            ax.xaxis.set_ticks_position("top")
            for s in ax.spines.values():
                s.set_visible(False)
            ax.tick_params(length=0)

            nfl = sum(1 for (p, _), _ in flagged.items() if p == plate)
            ax.set_title(f"{submission} — plate {plate}    "
                         f"{len(wells)} wells · {nfl} flagged for rerun",
                         fontsize=13, color=INK, pad=26, loc="left")

            sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
            cb = fig.colorbar(sm, ax=ax, fraction=.020, pad=.015)
            cb.set_label("MS1 TIC (×10⁹)", fontsize=8, color=INK_MUTED)
            # Ticks in the same ×10⁹ units the cells print, so the bar and the
            # numbers inside the wells can be read against each other. The
            # default offset text ("1e10") contradicts the axis label.
            cb.formatter = FuncFormatter(lambda v, _pos: f"{v / 1e9:.0f}")
            cb.update_ticks()
            cb.ax.tick_params(labelsize=7, colors=INK_MUTED)
            cb.outline.set_visible(False)

            # Legend: identity is never colour-alone.
            handles = [
                Rectangle((0, 0), 1, 1, fc="none", ec=CRITICAL, lw=2.4,
                          label="rerun — no data"),
                Rectangle((0, 0), 1, 1, fc="none", ec=WARNING, lw=2.4,
                          label="rerun — signal at blank level"),
                Rectangle((0, 0), 1, 1, fc=NEUTRAL, ec="#c9c7c2", hatch="///",
                          label="no analysis.tdf"),
            ]
            ax.legend(handles=handles, loc="upper left", bbox_to_anchor=(0, -.015),
                      ncol=3, frameon=False, fontsize=8, labelcolor=INK_MUTED)
            fig.tight_layout()
            pdf.savefig(fig, facecolor=SURFACE)
            plt.close(fig)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--scan", required=True, help="qc_scan.json from qc_scan.py")
    ap.add_argument("--submission", required=True, help="e.g. 793")
    ap.add_argument("--run-date", default=None,
                    help="date stamp for the rerun Sample IDs (default: today)")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--blank-frac", type=float, default=0.25,
                    help="fallback rule when --dashboard is not used: flag a sample\nbelow this fraction of its plate median")
    ap.add_argument("--dashboard", metavar="URL",
                    help="take STAN's needs_rerun as the authority on which acquired\nwells are suspect, e.g. https://ucd.stan-proteomics.org")
    ap.add_argument("--dash-token", help="share token for --dashboard")
    ap.add_argument("--instrument", default="timsTOF HT")
    ap.add_argument("--queue", action="append", default=[], metavar="XLSX",
                    help="submitted HyStar queue(s) to reconcile against; a queued\nwell with no directory at all is reported as never_ran (repeatable)")
    a = ap.parse_args()

    import datetime as dt
    run_date = a.run_date or dt.date.today().strftime("%Y%m%d")
    prot = f"PROT_{int(a.submission):04d}"
    out = Path(a.out_dir)
    out.mkdir(parents=True, exist_ok=True)

    rows = json.load(open(a.scan))
    dash = None
    if a.dashboard:
        dash = fetch_dashboard_reruns(a.dashboard, a.submission,
                                      a.dash_token, a.instrument)
        print(f"dashboard needs_rerun: {len(dash)} acquired wells flagged")
    reruns = find_reruns(rows, a.blank_frac, dash)
    if a.queue:
        queue = read_queue(a.queue)
        by_plate: dict[str, list[dict]] = {}
        for r in rows:
            by_plate.setdefault(r['plate'], []).append(r)
        never = []
        for grp in by_plate.values():
            never.extend(find_missing_wells(grp, queue))
        print(f'queue reconcile: {len(queue)} wells queued, {len(never)} never produced a directory')
        reruns = never + reruns
        reruns.sort(key=lambda r: (r['plate'], r['well'][0], int(r['well'][1:])))

    tsv = out / f"{prot}_reruns.tsv"
    xlsx = out / f"{prot}_rerun_queue.xlsx"
    pdf = out / f"{prot}_platemap.pdf"
    write_tsv(reruns, tsv)
    write_queue(reruns, xlsx, run_date, a.submission)
    write_platemap(rows, reruns, pdf, prot)

    from collections import Counter
    by = Counter(r["reason"] for r in reruns)
    src = "STAN dashboard" if dash is not None else f"local TIC < {a.blank_frac:.0%} of median"
    print(f"{prot}: {len(rows)} runs scanned, {len(reruns)} reruns "
          f"[{src}] — " + ", ".join(f"{v} {k}" for k, v in sorted(by.items())))
    for p in (tsv, xlsx, pdf):
        print(f"   {p}  ({p.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
